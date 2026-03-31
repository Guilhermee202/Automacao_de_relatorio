
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime, timedelta
from openpyxl.workbook.properties import CalcProperties


#  CONFIGURAÇÃO

MES_FIXO = 3     
ANO_FIXO = 2026

# Caminhos

caminho_analitico = r'P:\Relatóio\Analitico
caminho_retornos = r'P:\Relatório\Retorno


# Lista de meses
nomes_meses = ['JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO', 
               'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']

nome_mes_pt = nomes_meses[MES_FIXO - 1]

relatorio_destino = rf'V:RELATÓRIO\DESTINO\{ANO_FIXO}\{MES_FIXO} - {nome_mes_pt.capitalize()}\ICATU TRICARD SMS NPS_{nome_mes_pt}.xlsx'


#Geração de datas do mês

primeiro_dia_mes = datetime(ANO_FIXO, MES_FIXO, 1)
proximo_mes = primeiro_dia_mes.replace(day=28) + timedelta(days=4)
ultimo_dia_mes = proximo_mes - timedelta(days=proximo_mes.day)

datas = [primeiro_dia_mes + timedelta(days=i) for i in range((ultimo_dia_mes - primeiro_dia_mes).days + 1)]
datas_csv = [d.strftime('%Y%m%d') for d in datas]
datas_analitico = [d.strftime('%d%m%Y') for d in datas]


# Funções de carregamento

def carregar_analitico(data_ddmmaaaa):
    arquivos = [
        f'RELATORIO_ENVIO_SMS_{data_ddmmaaaa}_13H.csv',
        f'RELATORIO_ENVIO_SMS_{data_ddmmaaaa}_18H.csv'
    ]
    dfs = []
    for arquivo in arquivos:
        caminho = os.path.join(caminho_analitico, arquivo)
        if os.path.isfile(caminho):
            print(f'Lendo: {arquivo}')
            df = pd.read_csv(caminho, sep=';', encoding='latin1', dtype=str)
            if df.shape[1] > 5:
                df.drop(df.columns[5], axis=1, inplace=True)  # Remove coluna F
                df = df.iloc[:, 1:]  # Da coluna B em diante
                dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

def carregar_csv_padrao(prefixo, data_yyyymmdd):
    arquivo = f'{prefixo}_{data_yyyymmdd}.csv'
    caminho = os.path.join(caminho_retornos, arquivo)
    if os.path.isfile(caminho):
        print(f'Lendo: {arquivo}')
        return pd.read_csv(caminho, sep=';', encoding='latin1', dtype=str)
    return pd.DataFrame()


# Leitura e preparação da planilha

wb = load_workbook(relatorio_destino)
ws_analitico = wb['Analitico']
ws_ret_nota = wb['Retorno da Nota']
ws_retorno = wb['Retorno']

def limpar_colunas(ws, col_inicio, col_fim):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in range(col_inicio, col_fim + 1):
            row[col_idx - 1].value = None

limpar_colunas(ws_analitico, 2, 7)     # B até G
limpar_colunas(ws_ret_nota, 8, 14)     # H até N
limpar_colunas(ws_retorno, 1, 13)      # A até M


# Carrega os dados

df_analitico_total = pd.concat(
    [carregar_analitico(data) for data in datas_analitico if not carregar_analitico(data).empty],
    ignore_index=True
)

df_retnota_total = pd.concat(
    [carregar_csv_padrao('RETORNO_DA_NOTA', data) for data in datas_csv if not carregar_csv_padrao('RETORNO_DA_NOTA', data).empty],
    ignore_index=True
)

df_retorno_total = pd.concat(
    [carregar_csv_padrao('RETORNO', data) for data in datas_csv if not carregar_csv_padrao('RETORNO', data).empty],
    ignore_index=True
)

# Cola os dados nas abas

def colar_df(ws, df, start_col):
    start_row = 2
    for i, row in df.iterrows():
        for j, val in enumerate(row, start=start_col):
            ws.cell(row=start_row + i, column=j, value=val)

# Converte coluna L (coluna 12) da sheet Retorno da Nota
# Índice da coluna L no dataframe

coluna_L_idx = 12 - 8

if df_retnota_total.shape[1] > coluna_L_idx:
    
    col = df_retnota_total.iloc[:, coluna_L_idx].astype(str).str.strip()

    # Vazio vira traço
    col = col.replace(['', 'nan', 'None', 'NaN'], ' - ')

    # Zero continua número
    col = col.apply(lambda x: 0 if x == '0' else x)

    df_retnota_total.iloc[:, coluna_L_idx] = col


colar_df(ws_analitico, df_analitico_total, 2)
colar_df(ws_ret_nota, df_retnota_total, 8)
colar_df(ws_retorno, df_retorno_total, 1)

# Garante recálculo automático

try:
    if hasattr(wb, "calc_properties") and wb.calc_properties is not None:
        wb.calc_properties.fullCalcOnLoad = True
    else:
        wb.calc_properties = CalcProperties(fullCalcOnLoad=True)
except Exception as e:
    print(f"Recálculo automático falhou: {e}")

#  Salva 
wb.save(relatorio_destino)
wb.close()
print('Relatório atualizado com sucesso.')