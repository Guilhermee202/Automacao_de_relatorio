
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator

pasta_origem = r"P:\Caminho\Origem"
arquivo_excel = r"V:\Relatório
nome_sheet = "BASE CLIENTE"


# FUNÇÃO PARA PEGAR O ARQUIVO CSV MAIS RECENTE POR DATA NO NOME 
def pegar_csv_mais_recente(pasta):
    padrao = re.compile(r"RELATORIO_(\d{8})\.csv", re.IGNORECASE)
    arquivos = []

    for arq in os.listdir(pasta):
        match = padrao.match(arq)
        if match:
            data_str = match.group(1)
            try:
                data = pd.to_datetime(data_str, format="%Y%m%d")
                arquivos.append((arq, data))
            except:
                pass

    if not arquivos:
        raise FileNotFoundError(
            "Nenhum arquivo CSV com padrão RELATORIO_YYYYMMDD.csv encontrado."
        )

    arquivos.sort(key=lambda x: x[1], reverse=True)
    return arquivos[0][0]


# EXECUÇÃ

try:
    arquivo_csv = pegar_csv_mais_recente(pasta_origem)
    caminho_csv = os.path.join(pasta_origem, arquivo_csv)
    print(f"Arquivo CSV mais recente: {arquivo_csv}")
except Exception as e:
    print("Erro ao encontrar arquivo CSV:", e)
    exit(1)


# LER CSV
df = pd.read_csv(caminho_csv, sep=";", encoding="latin1", engine="python")


# ABRIR EXCEL
wb = load_workbook(arquivo_excel)

if nome_sheet not in wb.sheetnames:
    raise ValueError(f"A sheet '{nome_sheet}' não existe no arquivo Excel.")

ws = wb[nome_sheet]


# IDENTIFICAR ÚLTIMA LINHA ATUAL
ultima_linha_planilha = ws.max_row


#LIMPAR BASE ANTIGA
# Limpar A2 até Y
for linha in range(2, ultima_linha_planilha + 1):
    for col in range(1, 26):  # A até Y
        ws.cell(row=linha, column=col).value = None

# Limpar fórmulas antigas de Z3 até AK
for linha in range(3, ultima_linha_planilha + 1):
    for col in range(26, 38):  # Z até AK
        ws.cell(row=linha, column=col).value = None


# ESCREVER DADOS CSV (A2 até Y)
for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    for col_idx, valor in enumerate(row, start=1):
        if col_idx <= 25:  # garante que só escreve até Y
            ws.cell(row=row_idx, column=col_idx).value = valor


# DEFINIR ÚLTIMA LINHA COM DADOS
ultima_linha = len(df) + 1


# REPLICAR FÓRMULAS (Z até AK)
for col_idx in range(26, 38):  # Z até AK
    col_letter = get_column_letter(col_idx)
    source_cell = ws[f"{col_letter}2"]

    if source_cell.value is None:
        continue

    for row_idx in range(3, ultima_linha + 1):
        target_cell = ws[f"{col_letter}{row_idx}"]

        if isinstance(source_cell.value, str) and source_cell.value.startswith("="):
            formula = Translator(
                source_cell.value, origin=f"{col_letter}2"
            ).translate_formula(f"{col_letter}{row_idx}")

            target_cell.value = formula
        else:
            target_cell.value = source_cell.value

        if source_cell.has_style:
            target_cell._style = source_cell._style


# SALVAR
wb.save(arquivo_excel)

print(f"Atualização concluída! {len(df)} linhas coladas em '{nome_sheet}'.")